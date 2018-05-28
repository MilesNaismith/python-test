import requests
from openpyxl import Workbook
import csv
import re
wb = Workbook()
ws = wb.active
pattern = r'89[0-9]{9}'  # поиск мобильных номеров
api_key = "f45946ec-bc30-458c-bba1-5ebd22b0b970"

# Считываем лог астериска
Dubli = list()
with open("log.csv") as f:
    reader = csv.reader(f)
    for row in reader:
        Dubli.append(row[2])
for i in range(len(Dubli)):
    Dubli[i] = Dubli[i][1:]
Dubli.remove('rc')
zvonki = list()

# Избавляемся от дублей
numbers = set(Dubli)
numbers = list(numbers)

# Скачиваем информацию из api в формате json и формируем список с ответами
company = list()
for i in range(len(numbers)):
    zvonki.append(Dubli.count(numbers[i]))  # Количество звонков за день
    if re.match(pattern, numbers[i]) is None:
        number = str(numbers[i])
        api_url = "https://search-maps.yandex.ru/v1/?text=" + number + "&type=biz&lang=ru_RU&apikey=" + api_key
        res = requests.get(api_url)
        tabs = res.json()

        # Проверка на существование в справочнике
        if len(tabs['features']) > 0:
            CompanyName = tabs['features'][0]['properties']['CompanyMetaData']['name']
            company.append(CompanyName)
        else:
            company.append("Не найдено")
    else:
        company.append("мобильный телефон")

# Заполнение таблицы
ws['A1'] = "Номер телефона"
ws['B1'] = "Название организации"
ws['C1'] = "Количество звонков"
for i in range(len(numbers)):
    xla = 'A' + str(i + 2)
    xlb = 'B' + str(i + 2)
    xlc = 'C' + str(i + 2)
    ws[xla] = numbers[i]
    ws[xlb] = company[i]
    ws[xlc] = zvonki[i]
wb.save('spisok.xlsx')
